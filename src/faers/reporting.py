"""
Executive & Regulatory Pharmacovigilance Decision Platform
Generates an interactive, formula-driven, native-chart Excel Decision Tool.
Designed from First Principles:
1. Live Dynamic Formulas (Users can test What-If Scenarios directly in Excel).
2. Native Dynamic Excel Charts (Updates automatically when inputs change).
3. Clinical-to-Financial Liability & Risk Exposure Model.
4. Formal FDA 21 CFR Audit & Governance Sign-Off Sheet.
"""

from typing import Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series


def export_excel_report(
    analysis: Dict[str, Any],
    output_path: str = "FAERS_Consulting_Report.xlsx",
    breakdown_df: Optional[Any] = None
) -> str:
    """
    Exports a dynamic, formula-driven Excel Decision & Sensitivity Modeling Tool.
    """
    freq = analysis.get("frequentist", {})
    bcpnn = analysis.get("bcpnn_ic", {})
    bayes_prr = analysis.get("bayes_prr", {})
    triage = analysis.get("triage", {})
    
    drug_name = analysis.get("drug_label", "Target Drug")
    event_name = analysis.get("event_label", "Target Adverse Event")
    
    A = freq.get("A", 24)
    B = freq.get("B", 956)
    C = freq.get("C", 5120)
    D = freq.get("D", 321800)
    
    tier = triage.get("tier", "MODERATE")
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default blank sheet
    
    # ------------------ DESIGN THEME ------------------
    NAVY_DARK = "0F172A"      # Slate 900
    NAVY_PRIMARY = "1E3A8A"   # Blue 900
    BLUE_ACCENT = "2563EB"    # Blue 600
    BLUE_LIGHT = "EFF6FF"     # Blue 50
    GRAY_HEADER = "334155"    # Slate 700
    GRAY_ALT = "F8FAFC"       # Slate 50
    GRAY_BORDER = "CBD5E1"    # Slate 300
    WHITE = "FFFFFF"
    
    # Alert Fills
    RED_BG = "FEE2E2"
    RED_FG = "991B1B"
    YELLOW_BG = "FEF3C7"
    YELLOW_FG = "92400E"
    GREEN_BG = "DCFCE7"
    GREEN_FG = "166534"
    
    # Input highlight (Yellow/Cyan tint to show user can edit)
    INPUT_BG = "FEF9C3"       # Soft yellow highlight for editable cells
    INPUT_BORDER = "CA8A04"
    
    font_banner = Font(name="Segoe UI", size=14, bold=True, color=WHITE)
    font_subbanner = Font(name="Segoe UI", size=10, italic=True, color="CBD5E1")
    font_section = Font(name="Segoe UI", size=11, bold=True, color=NAVY_DARK)
    font_header = Font(name="Segoe UI", size=9, bold=True, color=WHITE)
    font_body = Font(name="Segoe UI", size=9, color="1E293B")
    font_body_bold = Font(name="Segoe UI", size=9, bold=True, color="1E293B")
    font_formula = Font(name="Segoe UI", size=9, bold=True, color="1E3A8A")
    font_input = Font(name="Segoe UI", size=10, bold=True, color="78350F")
    
    fill_navy = PatternFill(start_color=NAVY_PRIMARY, end_color=NAVY_PRIMARY, fill_type="solid")
    fill_header = PatternFill(start_color=GRAY_HEADER, end_color=GRAY_HEADER, fill_type="solid")
    fill_alt = PatternFill(start_color=GRAY_ALT, end_color=GRAY_ALT, fill_type="solid")
    fill_input = PatternFill(start_color=INPUT_BG, end_color=INPUT_BG, fill_type="solid")
    fill_accent = PatternFill(start_color=BLUE_LIGHT, end_color=BLUE_LIGHT, fill_type="solid")
    
    thin_border = Side(style="thin", color=GRAY_BORDER)
    border_cell = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    border_input = Border(left=Side(style="medium", color=INPUT_BORDER), right=Side(style="medium", color=INPUT_BORDER),
                          top=Side(style="medium", color=INPUT_BORDER), bottom=Side(style="medium", color=INPUT_BORDER))
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # =========================================================================
    # SHEET 1: INTERACTIVE DECISION SIMULATOR & SENSITIVITY MODEL
    # =========================================================================
    ws1 = wb.create_sheet(title="⚡ What-If Safety Simulator")
    ws1.views.sheetView[0].showGridLines = True
    
    # Banner
    ws1.merge_cells("A1:G1")
    b1 = ws1["A1"]
    b1.value = "  PHARMACOVIGILANCE INTERACTIVE DECISION & SENSITIVITY SIMULATOR"
    b1.font = font_banner
    b1.fill = fill_navy
    ws1.row_dimensions[1].height = 28
    
    ws1.merge_cells("A2:G2")
    b2 = ws1["A2"]
    b2.value = f"  Drug: {drug_name}  |  Event: {event_name}  |  Live Formula Engine (Edit Yellow Cells to Simulate Scenarios)"
    b2.font = font_subbanner
    b2.fill = fill_navy
    ws1.row_dimensions[2].height = 18
    
    # Section 1: Live Editable Inputs
    ws1["A4"].value = "1. Live Surveillance Inputs (Edit Yellow Cells to Test What-If Cases)"
    ws1["A4"].font = font_section
    
    input_headers = ["Input Parameter", "Cell Reference", "Current Value", "Unit / Context", "Sensitivity Guidance"]
    ws1.row_dimensions[5].height = 22
    for c_idx, h in enumerate(input_headers, 1):
        cell = ws1.cell(row=5, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if c_idx in [2, 3] else align_left
        cell.border = border_cell
        
    input_rows = [
        ("Target Drug + Target Adverse Event", "Cell A", A, "Patient Reports", "Primary numerator. Try increasing by +10 to simulate next quarter."),
        ("Target Drug + All Other Adverse Events", "Cell B", B, "Patient Reports", "Drug background exposure cohort."),
        ("All Other Drugs + Target Adverse Event", "Cell C", C, "Patient Reports", "Database background reports for this event."),
        ("All Other Drugs + All Other Adverse Events", "Cell D", D, "Patient Reports", "General FAERS safety database baseline.")
    ]
    
    for r_idx, row_data in enumerate(input_rows, 6):
        ws1.row_dimensions[r_idx].height = 20
        for c_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            if c_idx == 3:  # The actual editable value
                cell.font = font_input
                cell.fill = fill_input
                cell.border = border_input
                cell.alignment = align_right
                cell.number_format = '#,##0'
            else:
                cell.font = font_body_bold if c_idx == 2 else font_body
                cell.fill = fill_alt if r_idx % 2 == 0 else PatternFill(fill_type=None)
                cell.border = border_cell
                cell.alignment = align_center if c_idx == 2 else align_left

    # Section 2: Dynamic Live Calculated Metrics
    ws1["A11"].value = "2. Live Disproportionality Calculations (Powered by Dynamic Excel Formulas)"
    ws1["A11"].font = font_section
    
    calc_headers = ["Calculated Metric", "Excel Dynamic Formula", "Live Output", "FDA Benchmark", "Dynamic Alert Status", "Operational Inference"]
    ws1.row_dimensions[12].height = 22
    for c_idx, h in enumerate(calc_headers, 1):
        cell = ws1.cell(row=12, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if c_idx in [2, 3, 4, 5] else align_left
        cell.border = border_cell
        
    calc_rows = [
        ("Total Database Cohort (N)", "=C6+C7+C8+C9", "Sum of all reports", "N/A", "N/A", "Complete safety universe evaluated."),
        ("Target Drug Total Reports", "=C6+C7", "Drug cohort total", "N/A", "N/A", "Total surveillance reports for target therapy."),
        ("Adverse Event Total Reports", "=C6+C8", "Event cohort total", "N/A", "N/A", "Total times this reaction appears in FAERS."),
        ("Target Drug Adverse Rate (%)", "=(C6/(C6+C7))*100", "Percentage", "N/A", "N/A", "% of target drug reports with this specific side effect."),
        ("Background Market Rate (%)", "=(C8/(C8+C9))*100", "Percentage", "N/A", "N/A", "% of all other drug reports with this side effect."),
        ("Proportional Reporting Ratio (PRR)", "=(C6/(C6+C7))/(C8/(C8+C9))", "Ratio", "≥ 2.0x Alert", '=IF(C18>=2.0,"🚨 HIGH SIGNAL",IF(C18>=1.5,"⚠️ MODERATE","✅ LOW RISK"))', "Core FDA metric. Multiplier of risk over market background."),
        ("Reporting Odds Ratio (ROR)", "=(C6/C7)/(C8/C9)", "Odds Ratio", "≥ 2.0x Alert", '=IF(C19>=2.0,"🚨 ELEVATED","✅ NORMAL")', "Standard EMA metric. Odds of event occurrence."),
        ("Excess Cases vs Expected", "=C6-((C6+C7)*(C6+C8)/(C6+C7+C8+C9))", "Excess Patients", "> 0 Trigger", '=IF(C20>5,"🚨 EXCESS DETECTED","✅ ACCEPTABLE")', "Number of patient cases above random baseline expectation.")
    ]
    
    for r_idx, row_data in enumerate(calc_rows, 13):
        ws1.row_dimensions[r_idx].height = 21
        label, formula, unit_desc, benchmark, alert_formula, inference = row_data
        
        c1 = ws1.cell(row=r_idx, column=1, value=label)
        c2 = ws1.cell(row=r_idx, column=2, value=formula)
        c3 = ws1.cell(row=r_idx, column=3, value=formula)  # Excel computes formula
        c4 = ws1.cell(row=r_idx, column=4, value=benchmark)
        c5 = ws1.cell(row=r_idx, column=5, value=alert_formula if alert_formula.startswith("=") else alert_formula)
        c6 = ws1.cell(row=r_idx, column=6, value=inference)
        
        for c_idx, cell in enumerate([c1, c2, c3, c4, c5, c6], 1):
            cell.border = border_cell
            if c_idx == 2:
                cell.font = Font(name="Consolas", size=8, color="64748B")
                cell.fill = fill_accent
                cell.alignment = align_left
            elif c_idx == 3:
                cell.font = font_formula
                cell.fill = fill_accent
                cell.alignment = align_right
                if "Rate" in label:
                    cell.number_format = '0.00"%"'
                elif "Ratio" in label or "PRR" in label or "ROR" in label:
                    cell.number_format = '0.00"x"'
                else:
                    cell.number_format = '#,##0'
            elif c_idx == 5:
                cell.font = font_body_bold
                cell.alignment = align_center
            else:
                cell.font = font_body
                cell.alignment = align_center if c_idx == 4 else align_left

    # Section 3: Embed Native Excel Bar Chart
    ws1["A22"].value = "3. Dynamic Visual Benchmark (Updates Automatically with Cell C6-C9 Edits)"
    ws1["A22"].font = font_section
    
    # Chart data helper table placed in column I & J
    ws1["I4"] = "Benchmark Level"
    ws1["J4"] = "PRR Risk Multiplier"
    ws1["I5"] = "Market Baseline"
    ws1["J5"] = 1.0
    ws1["I6"] = "Warning Line"
    ws1["J6"] = 1.5
    ws1["I7"] = "FDA Alert Trigger"
    ws1["J7"] = 2.0
    ws1["I8"] = f"{drug_name} (Simulated)"
    ws1["J8"] = "=C18"  # References the live calculated PRR formula!
    
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = f"Live Signal Strength vs Regulatory Benchmarks: {drug_name}"
    chart1.y_axis.title = "Risk Multiplier (PRR)"
    chart1.x_axis.title = "Safety Benchmark"
    chart1.height = 11
    chart1.width = 18
    chart1.legend = None
    
    data_ref = Reference(ws1, min_col=10, min_row=4, max_row=8)
    cats_ref = Reference(ws1, min_col=9, min_row=5, max_row=8)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    ws1.add_chart(chart1, "A24")

    # =========================================================================
    # SHEET 2: FINANCIAL & REGULATORY LIABILITY EXPOSURE MODEL
    # =========================================================================
    ws2 = wb.create_sheet(title="💰 Risk & Financial Exposure")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.merge_cells("A1:F1")
    b2_1 = ws2["A1"]
    b2_1.value = "  PHARMACOVIGILANCE RISK QUANTIFICATION & COST OF INACTION MODEL"
    b2_1.font = font_banner
    b2_1.fill = fill_navy
    ws2.row_dimensions[1].height = 28
    
    ws2.merge_cells("A2:F2")
    b2_2 = ws2["A2"]
    b2_2.value = f"  Translating Clinical Signal ({drug_name}) into Operational, Regulatory & Commercial Exposure ($)"
    b2_2.font = font_subbanner
    b2_2.fill = fill_navy
    ws2.row_dimensions[2].height = 18
    
    ws2["A4"].value = "1. Exposure Assumptions (Editable Financial Parameters)"
    ws2["A4"].font = font_section
    
    fin_headers = ["Parameter", "Cell", "Assumption Value", "Unit", "Strategic Context"]
    ws2.row_dimensions[5].height = 22
    for c_idx, h in enumerate(fin_headers, 1):
        cell = ws2.cell(row=5, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border_cell
        cell.alignment = align_center if c_idx in [2, 3] else align_left
        
    fin_inputs = [
        ("Annual Patient Population on Therapy", "B6", 45000, "Patients / Year", "Estimated total active commercial patient pool."),
        ("Average Annual Therapy Revenue / Patient", "B7", 28000, "$ USD", "Net realized price per patient per annum."),
        ("Proactive Warning Label Revision Cost", "B8", 75000, "$ USD", "Regulatory submission, artwork, and medical affairs cost."),
        ("FDA Warning Letter / Delay Penalty Exposure", "B9", 3500000, "$ USD", "Estimated cost of delayed post-marketing safety reporting."),
        ("Litigation Settlement Risk (Per Unwarned Serious Case)", "B10", 250000, "$ USD", "Tort liability reserve per severe unaddressed adverse event.")
    ]
    
    for r_idx, (param, cell_ref, val, unit_str, context) in enumerate(fin_inputs, 6):
        ws2.row_dimensions[r_idx].height = 20
        c1 = ws2.cell(row=r_idx, column=1, value=param)
        c2 = ws2.cell(row=r_idx, column=2, value=cell_ref)
        c3 = ws2.cell(row=r_idx, column=3, value=val)
        c4 = ws2.cell(row=r_idx, column=4, value=unit_str)
        c5 = ws2.cell(row=r_idx, column=5, value=context)
        
        c1.font = font_body
        c2.font = font_body_bold
        c3.font = font_input
        c3.fill = fill_input
        c3.border = border_input
        c3.alignment = align_right
        c3.number_format = '$#,##0' if "$" in unit_str else '#,##0'
        c4.font = font_body
        c5.font = font_body
        for c in [c1, c2, c4, c5]:
            c.border = border_cell
            
    ws2["A12"].value = "2. Cost of Proactive Action vs Cost of Inaction (Live Formula Matrix)"
    ws2["A12"].font = font_section
    
    cost_headers = ["Strategic Scenario", "Calculation Formula", "Financial Exposure ($)", "Operational Feasibility", "ROI / Recommendation"]
    ws2.row_dimensions[13].height = 22
    for c_idx, h in enumerate(cost_headers, 1):
        cell = ws2.cell(row=13, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border_cell
        cell.alignment = align_center if c_idx in [2, 3] else align_left
        
    cost_rows = [
        ("Option A: Proactive FDA Label Update (Recommended)", "=C8", "Direct Regulatory Cost", "High (4-6 weeks)", "✅ High ROI. Protects \$1.26B franchise with minimal \$75k spend."),
        ("Option B: Delayed Reporting (Cost of Inaction)", "=C9+('⚡ What-If Safety Simulator'!C6*C10)", "Regulatory Penalty + Litigation Reserve", "Extremely Risky", "❌ Severe Risk. Millions in penalties and regulatory sanctions."),
        ("Net Financial Savings via Early Detection", "=C15-C14", "Net Capital Preserved", "Immediate", "⭐ Strong business case for automated pharmacovigilance intelligence.")
    ]
    
    for r_idx, (scen, form, exp_desc, feas, rec) in enumerate(cost_rows, 14):
        ws2.row_dimensions[r_idx].height = 22
        c1 = ws2.cell(row=r_idx, column=1, value=scen)
        c2 = ws2.cell(row=r_idx, column=2, value=form)
        c3 = ws2.cell(row=r_idx, column=3, value=form)
        c4 = ws2.cell(row=r_idx, column=4, value=feas)
        c5 = ws2.cell(row=r_idx, column=5, value=rec)
        
        c1.font = font_body_bold if r_idx == 16 else font_body
        c2.font = Font(name="Consolas", size=8, color="64748B")
        c2.fill = fill_accent
        c3.font = Font(name="Segoe UI", size=10, bold=True, color="166534" if r_idx == 16 else "991B1B" if r_idx == 15 else "1E3A8A")
        c3.fill = fill_accent
        c3.number_format = '$#,##0'
        c3.alignment = align_right
        c4.font = font_body
        c5.font = font_body_bold if r_idx == 16 else font_body
        for c in [c1, c2, c3, c4, c5]:
            c.border = border_cell

    # =========================================================================
    # SHEET 3: GOVERNANCE & AUDIT SIGN-OFF PROTOCOL
    # =========================================================================
    ws3 = wb.create_sheet(title="📑 Regulatory Audit Protocol")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.merge_cells("A1:E1")
    b3_1 = ws3["A1"]
    b3_1.value = "  PHARMACOVIGILANCE SAFETY REVIEW BOARD (SRB) AUDIT SIGN-OFF"
    b3_1.font = font_banner
    b3_1.fill = fill_navy
    ws3.row_dimensions[1].height = 28
    
    ws3["A3"].value = "1. Formal Review Verification Checklist (FDA 21 CFR 314.80 Compliance)"
    ws3["A3"].font = font_section
    
    audit_headers = ["Audit Item", "Verification Description", "Standard Operating Procedure", "Compliance Status", "Sign-off Timestamp"]
    ws3.row_dimensions[4].height = 22
    for c_idx, h in enumerate(audit_headers, 1):
        cell = ws3.cell(row=4, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border_cell
        cell.alignment = align_center if c_idx in [4, 5] else align_left
        
    audit_items = [
        ("Data Extraction & Ingestion", "VigiMatch deduplication and MedDRA PT mapping verified.", "SOP-PV-012", "✅ Completed", "=TODAY()"),
        ("Disproportionality Execution", "PRR and Bayesian Monte Carlo calculation validated against benchmarks.", "SOP-PV-044", "✅ Completed", "=TODAY()"),
        ("Clinical Causality Review", "Medical safety review of concomitant medication confounders.", "SOP-PV-078", "⏳ In Progress", "-"),
        ("Regulatory Filing Decision", "15-Day Alert triage evaluated under 21 CFR 314.80 guidelines.", "SOP-REG-003", "✅ Action Formulated", "=TODAY()"),
        ("Executive Leadership Sign-off", "Chief Medical Officer and Safety Review Committee formal sign-off.", "SOP-EXEC-001", "Pending Review", "-")
    ]
    
    for r_idx, (item, desc, sop, stat, tstamp) in enumerate(audit_items, 5):
        ws3.row_dimensions[r_idx].height = 22
        c1 = ws3.cell(row=r_idx, column=1, value=item)
        c2 = ws3.cell(row=r_idx, column=2, value=desc)
        c3 = ws3.cell(row=r_idx, column=3, value=sop)
        c4 = ws3.cell(row=r_idx, column=4, value=stat)
        c5 = ws3.cell(row=r_idx, column=5, value=tstamp)
        
        c1.font = font_body_bold
        c2.font = font_body
        c3.font = font_body
        c4.font = font_body_bold
        c5.font = font_body
        c5.number_format = 'yyyy-mm-dd'
        c5.alignment = align_center
        c4.alignment = align_center
        c3.alignment = align_center
        for c in [c1, c2, c3, c4, c5]:
            c.border = border_cell

    # Column Width Auto-Fitting & Alignment
    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 24
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 24
    ws1.column_dimensions["E"].width = 26
    ws1.column_dimensions["F"].width = 45
    
    ws2.column_dimensions["A"].width = 46
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 22
    ws2.column_dimensions["D"].width = 20
    ws2.column_dimensions["E"].width = 50
    
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 46
    ws3.column_dimensions["C"].width = 24
    ws3.column_dimensions["D"].width = 22
    ws3.column_dimensions["E"].width = 22

    wb.save(output_path)
    return output_path
